import os
from openpyxl import Workbook
from openpyxl import load_workbook

folder_path = "/home/chirayu/Downloads/champSim_delay_variation/"
txtfile_folder_path = folder_path + "results_4core_50M/"
string="CPU 3 cumulative IPC:"
def extract_val(current_file, wb, ws, string,key):
    fp = open(txtfile_folder_path+current_file, 'r')
    count=0
    line_number = [125]   #Actual line number - 1
    data_line = []
    for i, line in enumerate(fp):
        # data_line.append(line.strip())
        # if(count==1):
        #     data_line.append(line.strip)
        #     break

        if(line.find(string)!=-1):
            if(count==0):
                count=1

            else:
                data_line.append(line.strip)
                break
            

    # print(data_line)
  
    start_p = line.find(string) #Name of string
    if(start_p!=-1):
        start_p += 1+string.__len__()  #size of string + 1
        result = ''
        while(line[start_p] == ' '):
              start_p += 1
        while(line[start_p] != ' '):
              result += line[start_p]
              start_p += 1
    else:
        result=''
    return result

wb = Workbook()
ws = wb.active
ws.title = "Cumulative IPC"

filenames = os.listdir(txtfile_folder_path)

current_row = 1
cell_key = ws.cell(row = current_row, column = 1)
cell_val = ws.cell(row = current_row, column = 2)
cell_key.value = "Filename"
cell_val.value = "Cumulative IPC"
current_row += 1

for current_file in filenames:
    val = extract_val(current_file, wb, ws, string,key=1)
    cell_key = ws.cell(row = current_row, column = 1)
    cell_val = ws.cell(row = current_row, column = 2)
    cell_key.value = current_file
    cell_val.value = val
    current_row += 1

wb.save(filename = (folder_path + "custom_scripts/data_4core_test_with_warmup.xlsx"))
