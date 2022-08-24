import os
from openpyxl import Workbook
from openpyxl import load_workbook

folder_path = "/home/chirayu/Downloads/champSim_delay_variation/"
txtfile_folder_path = folder_path + "results_15M/"

def find_CUM_IPC(current_file, wb, ws):
    fp = open(txtfile_folder_path+current_file, 'r')
    line_number = [38]   #Actual line number - 1
    data_line = []
    for i, line in enumerate(fp):
        if i in line_number:
            data_line.append(line.strip())
        elif i > 60:
            break

    # print(data_line)
    for x in data_line:
        start_p = x.find("IPC") #Name of string
        start_p += 4    #size of string + 1
        result = ''
        while(x[start_p] == ' '):
            start_p += 1
        while(x[start_p] != ' '):
            result += x[start_p]
            start_p += 1
        return result

wb = Workbook()
ws = wb.active
ws.title = "Cumulative IPC"

filenames = os.listdir(txtfile_folder_path)

current_row = 1
cell_key = ws.cell(row = current_row, column = 1)
cell_val = ws.cell(row = current_row, column = 2)
cell_key.value = "Filename"
cell_val.value = "Cumulative IPC"
current_row += 1

for current_file in filenames:
    val = find_CUM_IPC(current_file, wb, ws)
    cell_key = ws.cell(row = current_row, column = 1)
    cell_val = ws.cell(row = current_row, column = 2)
    cell_key.value = current_file
    cell_val.value = val
    current_row += 1

wb.save(filename = (folder_path + "custom_scripts/data_15M.xlsx"))
